import django
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension

django.setup()
from collections import defaultdict
from django.views.decorators.csrf import csrf_exempt
from multiprocessing import Process, Manager
from django.shortcuts import render
from django.shortcuts import redirect
from django.http import JsonResponse
from django.http.response import HttpResponse, FileResponse
from .models import *
from zipfile import ZipFile
import time as current_time
import subprocess
import shutil
import json
import os
import csv
import glob


def get_pattern(list, start_keyword, end_keyword):
    result = []
    found = False
    for i in range(len(list) - 1, -1, -1):
        if list[i] == '': continue
        if list[i].split(',')[1] == end_keyword:
            while i >= 0:
                temp = list[i].split(',')[1]
                result.append(temp)
                if temp == start_keyword: found = True; break
                i -= 1
            break
    if found: return result
    return []


def get_iteration(list, start_keyword, end_keyword):
    result = []
    found = False
    for i in range(len(list) - 1, -1, -1):
        if list[i] == '': continue
        if end_keyword in list[i]:
            while i >= 0:
                result.append(list[i])
                if start_keyword in list[i]: found = True; break
                i -= 1
            break
    if found: return result
    return []


def module_start_check(s, config):
    for i in config["modules"]:
        if i[0] in s: return i
    return False


def module_end_check(s, config):
    for i in config["modules"]:
        if i[1] in s: return i
    return False


def get_time(s):
    tsi = s.index(';') + 1
    return int(s[tsi:tsi + 10])


def change_class_red(list_of_strings, index_of_string, index_of_class):
    l = list_of_strings
    i = index_of_string
    j = index_of_class
    l[i] = l[i][:j] + "red" + l[i][j + 3:]


def line_number(line):
    s = line
    return s[:s.find(")")]


def report_download(cd, config, p, test_name):
    c = 0
    d = 0
    e = 0
    test_name_current = test_name.split()[0]
    wb = Workbook()
    ws = wb.active
    red = PatternFill(fgColor="FF0000", fill_type="solid")
    thick_border = Border(left=Side(style='thick'),
                          right=Side(style='thick'),
                          top=Side(style='thick'),
                          bottom=Side(style='thick'))
    ft_h = Font(name='Open Sans', bold=False, size=11)
    ws.title = "Last Iteration " + test_name_current

    ws.append([test_name])
    ws.append(
        ['Module Name', 'Duration', 'Start Line Number', 'End Line Number', 'Start Time Stamp', 'End Time Stamp',
         'Start '
         'Syntax',
         'End Syntax', 'Inspection'])
    for i in config['modules']:
        li = []
        for line in p:
            if line.find(i[0]) != -1 and c == 0:
                li.append(line_number(line))
                li.append(get_time(line))
                li.append(line[39:])
                e = get_time(line)
                c = 1
                d = 0
            elif line.find(i[1]) != -1 and d == 0:
                li.append(line_number(line))
                li.append(get_time(line))
                li.append(line[39:])
                f = get_time(line)
                c = 0
                d = 1
                if f - e > i[3]:
                    ws.append([i[2], str(f - e), li[0], li[3], li[1], li[4], li[2], li[5], "Required"])
                else:
                    ws.append([i[2], str(f - e), li[0], li[3], li[1], li[4], li[2], li[5], "Not Required"])
                li.clear()

        for rows in ws.iter_rows():
            for cell in rows:
                cell.alignment = Alignment(wrapText=True)
                cell.font = ft_h
                if cell.value == "Required":
                    cell.fill = red
                cell.border = thick_border

        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=22)

        ws.column_dimensions = dim_holder
    wb.save(cd + test_name_current + '.xlsx')
    return test_name_current


# Create your views here.
def login(request):
    if 'username' in request.session:
        return redirect(index)
    if request.method == 'POST':
        un = request.POST["username"]
        pwd = request.POST["password"]
        try:
            obj = Users.objects.get(username=un)
            if obj.password == pwd:
                request.session['username'] = un
                request.session['type'] = obj.type
                return redirect(index)
            else:
                return render(request, "login.html", {"invalid_username": False, "invalid_password": True})
        except Exception as e:
            print(e)
            return render(request, "login.html", {"invalid_username": True, "invalid_password": False})
    return render(request, "login.html", {"invalid_username": False, "invalid_password": False})


def index(request):
    if 'username' in request.session:
        try:
            obj = ConfigFiles.objects.all()
        except Exception as e:
            return HttpResponse("Sorry, some error please try again...!\nThe Error is: " + str(e))
        return render(request, "index.html", {"files": obj, "user": request.session['username'].upper()})
    else:
        return redirect(login)


def logout(request):
    if 'username' in request.session:
        request.session.flush()
    return redirect(login)


def pat_for_one_testcase(rresponse: object, rreport: object, cd: object, zip_name: object, test_case: object, test_num: object, test_name: object) -> object:
    try:
        # print(zip_name)
        # newpath = r'C:\Program Files\arbitrary'
        # if not os.path.exists(newpath):
        #     os.makedirs(newpath)
        zip = zip_name[:-3]
        files = ["/Traces.log", "/Traces.log.bak"]
        files2 = ["/ws.log", "/ws.log.bak"]
        files3 = ["/cadd.log", "/cadd.log.bak"]
        inted = cd + "Integration Team logs/"
        r10team = cd + "R10 Team logs/"
        Hardwareteam = cd + "Hardware Team logs"
        cmd = ['7z', 'e', cd + zip + '.7z', '-o' + inted]

        print(cd, zip, cmd, inted)
        for i in files:
            print(i)
            cmd.append(zip + i)

        cmd2 = ['7z', 'e', cd + zip + '.7z', '-o' + r10team]

        print(cd, zip, cmd2, r10team)
        for i in files2:
            print(i)
            cmd2.append(zip + i)

        cmd3 = ['7z', 'e', cd + zip + '.7z', '-o' + Hardwareteam]

        print(cd, zip, cmd3, Hardwareteam)
        for i in files3:
            print(i)
            cmd3.append(zip + i)

        sp = subprocess.check_output(cmd)
        sp = subprocess.check_output(cmd2)
        sp = subprocess.check_output(cmd3)

        # files1 = ["/ws.log", "/ws.log.bak"]
        #
        # pd = pd + "R10 Team logs/"
        # cmd = ['7z', 'e', pd + zip + '.7z', '-o' + pd]
        #
        # print(pd, zip, cmd)
        # for i in files:
        #     print(i)
        #     cmd.append(zip + i)
        # sp = subprocess.check_output(cmd)

        list_files = [f for f in os.listdir(cd) if os.path.isfile(os.path.join(cd, f))]

        # print(list_files[1],list_files[2])

        traces = []
        if len(list_files) == 3:
            traces = open(cd + 'Traces.log.bak', 'rb').read().decode('utf-16').split('\n')

        traces += open(cd + 'Traces.log', 'rb').read().decode('utf-16').split('\n')

        # print(traces)

        start_time = 0
        end_time = len(traces)

        k = []
        with open(test_case, 'r') as json_data:
            config = json.load(json_data)
            for i in config['test_case_pattern']:
                k.extend(config['operations'][i])

        if test_name != "All Test Case":
            # k = []
            # with open(test_case, 'r') as json_data:
            #     config = json.load(json_data)
            #     for i in config['test_case_pattern']:
            #         k.extend(config['operations'][i])
            l = set(k)
            columns1 = defaultdict(list)
            j = []
            for line in traces:
                for i in l:
                    if i in line:
                        j.append(str(get_time(line)) + ',' + i + ',' + str(traces.index(line)) + '\n')

            test_case_scenario_length = 0

            # +EventHandler Event( ,Display,SmRestrictedItems,ChangeContext ) contains "," which separate two strings
            k = [sub.replace("+EventHandler Event( ,Display,WaitForApproval,ChangeContext )", "+EventHandler Event( ")
                 for sub in k]
            k = [sub.replace("+EventHandler Event( ,Display,SmRestrictedItems,ChangeContext )", "+EventHandler Event( ")
                 for sub in k]
            k = [sub.replace("+EventHandler Event( ,Display,ProduceFavorites0409,ChangeContext )",
                             "+EventHandler Event( ") for sub in k]
            k = [sub.replace("+EventHandler Event( ,TabSelected,ProduceFavorites0409,Click )", "+EventHandler Event( ")
                 for sub in k]
            k = [sub.replace("+EventHandler Event( ,PFKBLine3AlphaNumericKeys,ProduceFavorites0409,Click )",
                             "+EventHandler Event( ") for sub in k]

            reader = csv.reader(j)
            for row in reader:
                for (i, v) in enumerate(reader):
                    columns1[i].extend(v)

            # for i in range(len(columns1)):
            #     print(columns1[i][1])
            a = []
            c = []
            d = []
            e = 0
            for i in range(len(columns1)):
                if i + 1 < len(columns1):
                    if columns1[i][1] == k[0] and columns1[i + 1][1] == k[1]:
                        test_case_scenario_length = 2 * len(config["test_case_pattern"])
                        e = 1
                        c.clear()
                if e == 1:
                    test_case_scenario_length -= 1
                    if i + 1 < len(columns1):  # Update list of index out of range (error resolved)
                        if columns1[i][1] == k[0] and columns1[i + 1][1] == k[1]:
                            a.append(columns1[i][2])
                    c.append(columns1[i][1])
                if test_case_scenario_length == 0:
                    e = 0
                    if c == k:
                        d.append(a[-1])
                        d.append(columns1[i][2])
                    c.clear()

            if not d:
                # rresponse["test_case" + str(
                #     test_num)] = "<ul class=\"myUL\"><i class=\"fa fa-times\" style=\"padding:5px\"></i>Test case " + str(
                #     test_num) + " : Wrong Diag" + "</ul>"
                rreport.append([test_num, zip_name, test_name,
                                '<span style="color:red;font-family:poppins"><span class="fa fa-times"></span>Wrong '
                                'Diag</span>'])
                return None
            start_time = int(d[-2])
            end_time = int(d[-1])

        iteration = traces[start_time:end_time]
        test_name_current = report_download(cd, config, iteration, test_name)
        response = []
        parents = [0]
        index = 1
        d = dict()
        response.append(
            "<li><span class=\"caret not\">Test case " + str(
                test_num) + "&nbsp;&nbsp;" + test_name + "&nbsp;&nbsp;</span><a href=\"report?name=" + cd + test_name_current + ".xlsx\">&nbsp;<i class=\"fa fa-download\"></i></a><ul class=\"nested\">")
        for i in iteration:
            line = module_start_check(i, config)
            if line:
                d[line[0]] = [get_time(i), index]
                parents.append(index)
                response.append("<li><span class=\"caret not\">" + line[2])
                response.append("<ul class=\"nested\">")
                response.append("<li class=\"not\">" + i + "</li>")
                response.append("")
                index += 4
                continue
            line = module_end_check(i, config)
            if line:
                if line[0] in d:
                    time = get_time(i)
                    time_taken = time - d[line[0]][0]
                    start_line = d[line[0]][1]
                    response[start_line] += "( " + str(
                        time_taken) + "ms )</span>"  # default time(" + str(line[3]) + ")"
                    response[start_line + 3] = "<li class=\"not\">" + i + "</li>"
                    response.append("</ul>")
                    response.append("</li>")
                    index += 2
                    if time_taken > line[3]:
                        change_class_red(response, start_line + 2, 11)
                        change_class_red(response, start_line + 3, 11)
                        for i in parents:
                            change_class_red(response, i, 23)
                    parents.pop()
                    del d[line[0]]
                    continue
        response.append("</ul></li>")
        rresponse["test_case" + str(test_num)] = "<ul class=\"myUL\">" + "\n".join(response) + "</ul>"
        rreport.append([test_num, zip_name, test_name,
                        '<span style="color:#54b948;font-family:poppins"><span class="fa fa-check"></span>Correct Diag</span>'])
    except Exception as e:
        print(e)
        test_num = "You can now go to the logs collecting location to check different teams logs and errors!!!"
        if test_name == 'Customer Mode':
            rresponse[str(test_num)] = "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i>" + str(
                test_num) + "</ul><br>" + "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "Mode Selected " + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#ff0000\"> " + str(test_name) + "</b></FONT></ul><br>" +"<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "Type of Error " + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#ff0000\"> Environment Setup Issue" + "</b></FONT></ul><br>" + "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "Steps to Resolve it" + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#0000ff\"> Please install CADD_Full_Update_v1.9+ Patch_v3.4 or above!!!" + "</b></FONT></ul><br>"
        else:
            print(test_name)
            rresponse[str(test_num)] = "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i>" + str(
                test_num) + "</ul><br>" + "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "Mode Selected " + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#ff0000\"> " + str(test_name) + "</b></FONT></ul><br>" + "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "ADK Version " + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#32ACCB\"> SelfServ Checkout CoreApp_v6.2.0_b041" + "</b></FONT></ul><br>" + \
                                       "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "NGUI Version " + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#8B0000\"> NGUI_v1.2.5_b038_Hotfix_v1" + "</b></FONT></ul><br>" + \
                                       "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "Integration Build Number " + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#006400\"> Coles Lane Full 06.2.5.132.D7.1.0185" + "</b></FONT></ul><br>" + \
                                       "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "R10 Version " + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#4B0082\"> Drop 54 " + "</b></FONT></ul><br>" + \
                                       "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "TB Version " + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#F09511\"> TB 5.0 Runtime Package_v4.8.0.0022" + "</b></FONT></ul><br>" + \
                                       "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "Error Found in Logs " + "</u>" + " : &nbsp;&nbsp;<FONT COLOR=\"#F011E5\"> Device returns error = 10000" + "</b></FONT></ul><br>" + \
                                       "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "Team Resolved this Issue" + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#006400\"> Integration Team" + "</b></FONT></ul><br>" + "</ul><br>" + "<ul class=\"myUL\"><i class=\"fa fa-wrong\"></i><b><u>" + "Steps Recommended to Customer" + "</u>" + ": &nbsp;&nbsp;<FONT COLOR=\"#0000ff\"> Please install CADD_Full_Update_v1.9+ Patch_v3.4 or above!!!" + "</b></FONT></ul><br> "


def uploadfile(request):
    if 'username' in request.session:
        try:
            if __name__ == "PAT.views":
                start_time_float = current_time.time()
                start_time = request.session["username"] + str(start_time_float)
                cd = "./PAT/files/" + start_time
                obj = ConfigFiles.objects.all()
                testcase_dict = {}
                for i in obj:
                    testcase_dict[i.file.name] = i.file_name
                response = Manager().dict()
                rreport = Manager().list()
                processes = []
                j = 0
                for i in request.FILES:
                    j += 1
                    ZipFile(request.FILES[i]).extractall(cd + "/" + "test_case" + str(j) + "/")
                    mypath = cd + "/" + "test_case" + str(j) + "/"
                    f = os.listdir(mypath)
                    a = str(f[0])
                    processes.append(Process(target=pat_for_one_testcase, args=(
                        response, rreport, cd + "/" + "test_case" + str(j) + "/", a,
                        request.POST["test_case" + str(j)], j, testcase_dict[request.POST["test_case" + str(j)]])))
                    processes[j - 1].start()
                for i in processes:
                    i.join()

                request.session["test_cases"] = response.values()
                request.session["time_taken"] = round(current_time.time() - start_time_float, 3)
                request.session["report"] = sorted(rreport, key=lambda x: x[0])
                return HttpResponse("got it")

        except Exception as e:
            print(e)
            return HttpResponse("Sorry, some error please try again...!\nThe Error is: " + str(e))
    else:
        return redirect(login)


def addupdate(request):
    if 'username' in request.session:
        try:
            obj = ConfigFiles.objects.all()
            return render(request, "add-update.html", {"files": obj})
        except Exception as e:
            print(e)
            return HttpResponse("Sorry, some error please try again...!\nThe Error is: " + str(e))
    return redirect(login)


def addfile(request):
    if 'username' in request.session:
        try:
            obj = ConfigFiles()
            obj.file_name = request.POST["file_name"]
            obj.file = request.FILES["file"]
            obj.save()
            return redirect(addupdate)
        except Exception as e:
            print(e)
            os.remove("./ConfigFiles/" + request.FILES["file"].name)
            return HttpResponse("Sorry, some error please try again...!\nThe Error is: " + str(e))

    return redirect(login)


def add(request):
    if 'username' in request.session:
        return render(request, "add.html")
    return redirect(login)


def delete(request):
    if 'username' in request.session:
        obj = ConfigFiles.objects.get(id=request.GET["id"])
        os.remove(obj.file.name)
        obj.delete()
        return redirect(addupdate)
    return redirect(login)


def update(request):
    if 'username' in request.session:
        obj = ConfigFiles.objects.get(id=request.GET["id"])
        return render(request, "update.html", {"file_name": obj.file_name, "file": obj.file, 'id': obj.id})
    return redirect(login)


def updatefile(request):
    if 'username' in request.session:
        try:
            obj = ConfigFiles.objects.get(id=request.POST["id"])
            obj.file_name = request.POST["file_name"]
            try:
                old_file = obj.file
                obj.file = request.FILES["file"]
                print(obj.file.name, old_file.name)
                os.remove(old_file.name)
            except:
                pass
            obj.save()
            return redirect(addupdate)
        except Exception as e:
            print(e)
            return HttpResponse("Sorry, some error please try again...!\nThe Error is: " + str(e))
    return redirect(login)


def configfile(request):
    if 'username' in request.session:
        obj = ConfigFiles.objects.get(id=request.GET["id"])
        file = open(obj.file.name, 'rb')
        return FileResponse(file, content_type='application/json')
    return redirect(login)


def report(request):
    if 'username' in request.session:
        try:
            name = request.GET["name"]
            file = open(name, 'rb')
            return FileResponse(file, content_type='application/csv')
        except Exception as e:
            print(e)
            return HttpResponse("Sorry, some error please try again...!\nThe Error is: " + str(e))
    return redirect(login)


def pat(request):
    try:
        obj = ConfigFiles.objects.all()
    except Exception as e:
        return HttpResponse("Sorry, some error please try again...!\nThe Error is: " + str(e))
    return render(request, "pat.html", {"files": obj})


def output(request):
    if 'username' in request.session:
        try:
            test_cases = request.session["test_cases"]
            time_taken = request.session["time_taken"]
            report = request.session["report"]
            return render(request, "output.html",
                          {"test_cases": test_cases, "time_taken": time_taken, "report": report})
        except Exception as e:
            print(e)
            return HttpResponse("Sorry, some error please try again...!\nThe Error is: " + str(e))
    return redirect(login)


def test(request):
    return HttpResponse("iuh")
